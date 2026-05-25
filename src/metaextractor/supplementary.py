"""Supplementary materials fetcher and table/text extractor.

Resolves supplementary files for a PMCID through a layered chain of sources:
  1. PMC AWS S3 open-data bucket (canonical, anonymous, no anti-bot gating),
     keyed by the filenames discovered in the JATS XML.
  2. The Europe PMC supplementaryFiles ZIP bundle (fallback).

Per file:
  - xlsx/csv/tsv → parsed into one ``Table`` per sheet (xlsx) or per file.
  - pdf         → tables extracted with pdfplumber when available; remaining
                  page text returned as prose. Falls back to pypdf if
                  pdfplumber isn't installed.
  - txt         → prose only.
  - images / unsupported extensions → skipped with a recorded reason.

Each ``Supplementary`` carries both a flattened ``text`` representation
(retained so the existing LLM path keeps working until Phase 4) and a
structured ``tables`` list that downstream stages can consume directly
without re-parsing.
"""
from __future__ import annotations

import csv
import io
import urllib.request
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

EUROPEPMC_SUPP = "https://www.ebi.ac.uk/europepmc/webservices/rest/{pmcid}/supplementaryFiles"
PMC_S3_LIST = "https://pmc-oa-opendata.s3.amazonaws.com/?list-type=2&prefix={prefix}&delimiter=/"
PMC_S3_FILE = "https://pmc-oa-opendata.s3.amazonaws.com/{key}"
S3_NS = "{http://s3.amazonaws.com/doc/2006-03-01/}"
USER_AGENT = "metaextractor/0.1 (+https://github.com/OmicsMLRepo)"

TEXT_EXTS = {".csv", ".tsv", ".txt"}
SKIP_EXTS = {".eps", ".gif", ".jpg", ".jpeg", ".png", ".tif", ".tiff", ".svg"}


@dataclass
class Table:
    """A structurally-parsed table from one supplementary file/sheet.

    ``raw_rows`` preserves the unprocessed 2-D cell grid (cleaned of fully
    empty rows, cells coerced to stripped strings). Downstream stages that
    need to re-interpret the table — e.g. an LLM plan executor that wants
    to treat several leading rows as a stacked header rather than just
    row 0 — operate on ``raw_rows``. ``columns``/``rows`` carry the
    default "row-0 is header" interpretation.
    """

    name: str  # filename, optionally with ``::sheet`` suffix for xlsx
    source: str  # "s3" | "europepmc" | "local"
    columns: list[str]
    rows: list[dict[str, Any]]
    raw_rows: list[list[str]] = field(default_factory=list)


@dataclass
class Supplementary:
    """Result of resolving a paper's supplementary materials.

    ``text`` carries only prose content (from txt files and PDF page text
    outside of detected tables) — it never duplicates table contents, so
    the LLM never re-reads what the deterministic path will consume from
    ``tables``. ``tables`` holds the structured rows for downstream
    column-mapping + joining.
    """

    text: str
    tables: list[Table] = field(default_factory=list)
    included: list[str] = field(default_factory=list)
    skipped: list[tuple[str, str]] = field(default_factory=list)  # (filename, reason)


def _http_get(url: str, timeout: float = 60.0) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return resp.read()


def _rows_to_table(name: str, source: str, raw_rows: list[list[str]]) -> Table | None:
    """Promote a 2-D cell grid into a ``Table``.

    The first non-empty row is treated as the header. Returns None when the
    grid has fewer than two non-empty rows (no header + data row pair) or
    when the header is degenerate (zero non-empty columns).
    """
    cleaned = [
        [("" if c is None else str(c)).strip() for c in row]
        for row in raw_rows
        if any(("" if c is None else str(c)).strip() for c in row)
    ]
    if len(cleaned) < 2:
        return None
    max_width = max(len(r) for r in cleaned)
    cleaned = [r + [""] * (max_width - len(r)) for r in cleaned]
    header = cleaned[0]
    if not any(header):
        return None
    width = len(header)
    seen: dict[str, int] = {}
    columns: list[str] = []
    for i, h in enumerate(header):
        col = h or f"col_{i}"
        if col in seen:
            seen[col] += 1
            col = f"{col}_{seen[col]}"
        else:
            seen[col] = 0
        columns.append(col)
    rows: list[dict[str, Any]] = []
    for row in cleaned[1:]:
        row = row[:width] + [""] * max(0, width - len(row))
        rows.append({columns[i]: row[i] for i in range(width)})
    return Table(name=name, source=source, columns=columns, rows=rows, raw_rows=cleaned)


def _xlsx_to_tables(name: str, source: str, blob: bytes) -> list[Table]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError(
            "xlsx supplementary support requires the 'supplementary' extra: "
            "pip install 'metaextractor[supplementary]'"
        ) from e
    wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    out: list[Table] = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        raw_rows = [list(row) for row in ws.iter_rows(values_only=True)]
        tname = f"{name}::{sn}" if len(wb.sheetnames) > 1 else name
        t = _rows_to_table(tname, source, raw_rows)
        if t is not None:
            out.append(t)
    return out


def _csv_to_tables(name: str, source: str, blob: bytes, delimiter: str) -> list[Table]:
    text = blob.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    raw_rows = [row for row in reader]
    t = _rows_to_table(name, source, raw_rows)
    return [t] if t is not None else []


def _pdf_to_tables_and_text(name: str, source: str, blob: bytes) -> tuple[list[Table], str]:
    """Extract tables via pdfplumber when available; return (tables, prose).

    Tables found by pdfplumber are returned structurally and *not* duplicated
    into the prose text. When pdfplumber isn't installed (or the PDF yields
    no tables), falls back to pypdf for whole-page text.
    """
    try:
        import pdfplumber  # type: ignore[import-not-found]
    except ImportError:
        pdfplumber = None  # type: ignore[assignment]

    tables: list[Table] = []
    text_parts: list[str] = []

    if pdfplumber is not None:
        with pdfplumber.open(io.BytesIO(blob)) as pdf:
            for page_idx, page in enumerate(pdf.pages, start=1):
                page_tables = page.extract_tables() or []
                table_bboxes = []
                for ti, raw in enumerate(page_tables, start=1):
                    t = _rows_to_table(f"{name}::p{page_idx}_t{ti}", source, raw)
                    if t is not None:
                        tables.append(t)
                # Page text with table regions excluded (best-effort: pdfplumber
                # doesn't always give bboxes via extract_tables, so we use the
                # `find_tables` API for region exclusion).
                try:
                    finds = page.find_tables() or []
                    table_bboxes = [tb.bbox for tb in finds]
                except Exception:
                    table_bboxes = []
                if table_bboxes:
                    try:
                        non_table = page.filter(
                            lambda obj, bboxes=table_bboxes: not any(
                                _bbox_contains(b, obj) for b in bboxes
                            )
                        )
                        ptxt = non_table.extract_text() or ""
                    except Exception:
                        ptxt = page.extract_text() or ""
                else:
                    ptxt = page.extract_text() or ""
                if ptxt.strip():
                    text_parts.append(ptxt)
        return tables, "\n\n".join(text_parts)

    try:
        from pypdf import PdfReader
    except ImportError as e:
        raise RuntimeError(
            "pdf supplementary support requires the 'pdf' extra "
            "(pip install 'metaextractor[pdf]') or the 'supplementary' extra "
            "for table-aware parsing (adds pdfplumber)."
        ) from e
    reader = PdfReader(io.BytesIO(blob))
    return [], "\n\n".join(p.extract_text() or "" for p in reader.pages)


def _bbox_contains(bbox: tuple[float, float, float, float], obj: dict) -> bool:
    """True when an object's centroid sits inside a bbox (x0, top, x1, bottom)."""
    x0, top, x1, bottom = bbox
    cx = (obj.get("x0", 0) + obj.get("x1", 0)) / 2
    cy = (obj.get("top", 0) + obj.get("bottom", 0)) / 2
    return x0 <= cx <= x1 and top <= cy <= bottom


def _file_to_parts(
    name: str, source: str, blob: bytes
) -> tuple[list[Table], str]:
    """Convert one file's bytes to (tables, prose_text).

    Raises ValueError when the extension is image-only or unsupported.
    Lower-level converters may raise their own exceptions on parse failure.
    """
    ext = ("." + name.rsplit(".", 1)[-1].lower()) if "." in name else ""
    if ext in SKIP_EXTS:
        raise ValueError(f"image-only ({ext})")
    if ext == ".xlsx":
        return _xlsx_to_tables(name, source, blob), ""
    if ext == ".csv":
        return _csv_to_tables(name, source, blob, ","), ""
    if ext == ".tsv":
        return _csv_to_tables(name, source, blob, "\t"), ""
    if ext == ".pdf":
        return _pdf_to_tables_and_text(name, source, blob)
    if ext in TEXT_EXTS:
        return [], blob.decode("utf-8", errors="replace")
    raise ValueError(f"unsupported extension ({ext or '?'})")


def _block(name: str, source: str, text: str) -> str:
    return f"--- SUPPLEMENTARY FILE: {name} (source: {source}) ---\n{text.strip()}"


def _latest_s3_version_prefix(bare: str) -> str | None:
    """Return the highest-numbered version prefix (e.g. ``PMC5264247.1``)
    available in the PMC AWS S3 open-data bucket, or None if absent.
    """
    url = PMC_S3_LIST.format(prefix=f"PMC{bare}.")
    try:
        body = _http_get(url)
    except Exception:
        return None
    try:
        root = ET.fromstring(body)
    except ET.ParseError:
        return None
    versions: list[tuple[int, str]] = []
    for cp in root.findall(f"{S3_NS}CommonPrefixes/{S3_NS}Prefix"):
        if cp.text is None:
            continue
        head = cp.text.rstrip("/")
        if not head.startswith(f"PMC{bare}."):
            continue
        tail = head.rsplit(".", 1)[1]
        if tail.isdigit():
            versions.append((int(tail), head))
    if not versions:
        return None
    return max(versions)[1]


def fetch_supplementary(
    pmcid: str,
    jats_hrefs: list[tuple[str, str]] | None = None,
) -> Supplementary:
    """Resolve supplementary files for a PMCID through a layered source chain.

    Sources, in order:
      1. PMC AWS S3 (``pmc-oa-opendata``), keyed by filenames discovered in
         the JATS XML. Open, anonymous, no anti-bot gating.
      2. Europe PMC's supplementaryFiles ZIP for the article.

    Files are deduplicated by lowercased basename; later sources do not
    re-add filenames an earlier source already supplied.
    """
    bare = pmcid.upper().removeprefix("PMC")
    blocks: list[str] = []
    tables: list[Table] = []
    included: list[str] = []
    skipped: list[tuple[str, str]] = []
    seen: set[str] = set()

    # Source 1: PMC AWS S3, addressed by JATS-declared filenames.
    if jats_hrefs:
        version_prefix = _latest_s3_version_prefix(bare)
        if version_prefix is None and jats_hrefs:
            skipped.append((f"PMC{bare}", "s3: no version prefix found in pmc-oa-opendata"))
        elif version_prefix is not None:
            for name, _orig_url in jats_hrefs:
                key = name.lower()
                if key in seen:
                    continue
                ext = ("." + name.rsplit(".", 1)[-1].lower()) if "." in name else ""
                if ext in SKIP_EXTS:
                    skipped.append((name, f"image-only ({ext})"))
                    seen.add(key)
                    continue
                s3_url = PMC_S3_FILE.format(key=f"{version_prefix}/{name}")
                try:
                    blob = _http_get(s3_url)
                    file_tables, prose = _file_to_parts(name, "s3", blob)
                except Exception as e:
                    skipped.append((name, f"s3: {type(e).__name__}: {e}"))
                    continue
                tables.extend(file_tables)
                if prose.strip():
                    blocks.append(_block(name, "s3", prose))
                included.append(name)
                seen.add(key)

    # Source 2: Europe PMC supplementaryFiles ZIP.
    europepmc_url = EUROPEPMC_SUPP.format(pmcid=f"PMC{bare}")
    zf: zipfile.ZipFile | None = None
    try:
        zip_bytes = _http_get(europepmc_url)
        zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    except zipfile.BadZipFile:
        skipped.append(("(europepmc zip)", "response was not a valid ZIP"))
    except Exception as e:
        # Only surface the network failure when nothing else worked, so a
        # successful S3 pass isn't drowned out by a routine EuropePMC 404.
        if not included:
            skipped.append(("(europepmc download)", f"{type(e).__name__}: {e}"))

    if zf is not None:
        for name in sorted(zf.namelist()):
            base = name.rsplit("/", 1)[-1]
            key = base.lower()
            if key in seen:
                continue
            try:
                blob = zf.read(name)
                file_tables, prose = _file_to_parts(base, "europepmc", blob)
            except Exception as e:
                skipped.append((name, f"europepmc: {type(e).__name__}: {e}"))
                continue
            tables.extend(file_tables)
            if prose.strip():
                blocks.append(_block(base, "europepmc", prose))
            included.append(name)
            seen.add(key)

    return Supplementary(
        text="\n\n".join(blocks),
        tables=tables,
        included=included,
        skipped=skipped,
    )


def supplementary_from_local(paths: list[Path]) -> Supplementary:
    """Build a Supplementary record from user-supplied local files."""
    blocks: list[str] = []
    tables: list[Table] = []
    included: list[str] = []
    skipped: list[tuple[str, str]] = []
    for path in paths:
        try:
            blob = path.read_bytes()
            file_tables, prose = _file_to_parts(path.name, "local", blob)
        except Exception as e:
            skipped.append((str(path), f"local: {type(e).__name__}: {e}"))
            continue
        tables.extend(file_tables)
        if prose.strip():
            blocks.append(_block(path.name, "local", prose))
        included.append(str(path))
    return Supplementary(
        text="\n\n".join(blocks),
        tables=tables,
        included=included,
        skipped=skipped,
    )
